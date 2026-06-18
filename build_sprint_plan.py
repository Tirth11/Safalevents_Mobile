# -*- coding: utf-8 -*-
"""Generate the SafalEvents MOBILE Sprint Planning workbook (.xlsx).

Hierarchy: Epic -> Feature -> User Story, with developer assignment, sprint
timeline and the app/preview URL. Shares story data with the mobile user-story doc."""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from build_user_stories import EPICS  # single source of truth (no side effects on import)

OUT = r"E:\Safalvir\SafalEvents-Mobile\SafalEvents_Mobile_Sprint_Plan.xlsx"
# The mobile app also runs on the web (Expo / react-native-web) and is hosted from
# this repo. Replace with your Netlify web-preview URL once deployed.
MOCKUP_URL = "https://github.com/Tirth11/Safalevents_Mobile"
MOCKUP_DISPLAY = "Mobile app (Expo) — " + MOCKUP_URL

FEATURES = {
 "NAV": [
   ("FEAT-NAV-01", "Splash & role-based navigation", ["US-NAV-001", "US-NAV-002"]),
   ("FEAT-NAV-02", "Phone-frame, theming & live store", ["US-NAV-003", "US-NAV-004", "US-NAV-005"]),
 ],
 "BROWSE": [
   ("FEAT-BROWSE-01", "Guest Mode browse & discovery", ["US-BROWSE-001", "US-BROWSE-002", "US-BROWSE-003", "US-BROWSE-004", "US-BROWSE-009"]),
   ("FEAT-BROWSE-02", "Persistent entry & gated auth", ["US-BROWSE-005", "US-BROWSE-006"]),
   ("FEAT-BROWSE-03", "Return-to-intent, cancel & sessions", ["US-BROWSE-007", "US-BROWSE-008", "US-BROWSE-010"]),
 ],
 "AUTH": [
   ("FEAT-AUTH-01", "Signup (Guest/Host, Individual & Organization) + OTP", ["US-AUTH-001", "US-AUTH-002", "US-AUTH-003", "US-AUTH-004"]),
   ("FEAT-AUTH-02", "Login, staff invite login & session", ["US-AUTH-005", "US-AUTH-006", "US-AUTH-007"]),
 ],
 "GUEST": [
   ("FEAT-GUEST-01", "RSVP flow & tickets/QR pass", ["US-GUEST-001", "US-GUEST-002", "US-GUEST-003", "US-GUEST-004"]),
   ("FEAT-GUEST-02", "Guest messaging & profile", ["US-GUEST-005", "US-GUEST-006", "US-GUEST-007"]),
 ],
 "HOST": [
   ("FEAT-HOST-01", "Host dashboard & events list", ["US-HOST-001", "US-HOST-002"]),
   ("FEAT-HOST-02", "Create & manage events", ["US-HOST-003", "US-HOST-004", "US-HOST-005"]),
   ("FEAT-HOST-03", "Host account", ["US-HOST-006"]),
 ],
 "APPROVAL": [
   ("FEAT-APPROVAL-01", "Approval queue & waitlist", ["US-APPROVAL-001", "US-APPROVAL-002"]),
   ("FEAT-APPROVAL-02", "Status badges & live reflection", ["US-APPROVAL-003", "US-APPROVAL-004"]),
 ],
 "STAFF": [
   ("FEAT-STAFF-01", "Staff login & permission gating", ["US-STAFF-001", "US-STAFF-005"]),
   ("FEAT-STAFF-02", "QR gate check-in", ["US-STAFF-002", "US-STAFF-003"]),
   ("FEAT-STAFF-03", "Staff & roles management", ["US-STAFF-004"]),
 ],
 "MSG": [
   ("FEAT-MSG-01", "Per-event messaging config & threads", ["US-MSG-001", "US-MSG-002", "US-MSG-003"]),
 ],
 "PLAT": [
   ("FEAT-PLAT-01", "Aligned layout & web build", ["US-PLAT-001", "US-PLAT-002"]),
   ("FEAT-PLAT-02", "Deployment, persistence & prototype data", ["US-PLAT-003", "US-PLAT-004", "US-PLAT-005"]),
 ],
}

EPIC_SPRINT = {
 "NAV": 1, "PLAT": 1,
 "BROWSE": 2,
 "AUTH": 3,
 "GUEST": 4, "MSG": 4,
 "HOST": 5, "APPROVAL": 5,
 "STAFF": 6,
}
SPRINT_START = datetime.date(2026, 6, 22)
def sprint_dates(n):
    start = SPRINT_START + datetime.timedelta(days=(n - 1) * 14)
    return start, start + datetime.timedelta(days=11)
def timeline_str(n):
    s, e = sprint_dates(n)
    return "Sprint %d  ·  %s – %s" % (n, s.strftime("%d %b %Y"), e.strftime("%d %b %Y"))

DEVS = ["Developer 1", "Developer 2", "Developer 3", "Developer 4"]

story_by_id, epic_name = {}, {}
for ep in EPICS:
    epic_name[ep["id"]] = ep["name"]
    for s in ep["stories"]:
        story_by_id[s["id"]] = s
def story_desc(s):
    return "As a %s, I want to %s, so that %s." % (s["role"], s["want"], s["benefit"])

wb = Workbook(); ws = wb.active; ws.title = "Sprint Plan"
HEAD_FILL = PatternFill("solid", fgColor="F2541B")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(vertical="top", wrap_text=True); TOP = Alignment(vertical="top")
thin = Side(style="thin", color="E2D5CE"); BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
PRI_FILL = {"High": "FDE2E1", "Medium": "FFF3D6", "Low": "E3F0E6"}

COLS = [("Epic ID", 12), ("Epic Description", 28), ("Feature ID", 18), ("Feature Description", 34),
        ("User Story ID", 16), ("User Story Description", 62), ("Priority", 10),
        ("Developer Assigned", 16), ("Sprint", 9), ("Timeline", 30), ("Mockup URL", 40)]
for i, (h, w) in enumerate(COLS, start=1):
    c = ws.cell(row=1, column=i, value=h)
    c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = Alignment(vertical="center", wrap_text=True); c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 26; ws.freeze_panes = "A2"

r = 2; feat_counter = 0; rows_written = 0
for ep in EPICS:
    eid = ep["id"]; sprint_no = EPIC_SPRINT.get(eid, 6); tline = timeline_str(sprint_no)
    feats = list(FEATURES.get(eid, []))
    mapped = {sid for _, _, sids in feats for sid in sids}
    leftover = [s["id"] for s in ep["stories"] if s["id"] not in mapped]
    if leftover:
        feats.append(("FEAT-%s-99" % eid, "%s — additional stories" % ep["name"], leftover))
    for fid, fdesc, sids in feats:
        dev = DEVS[feat_counter % len(DEVS)]; feat_counter += 1
        for sid in sids:
            s = story_by_id.get(sid)
            if not s:
                continue
            vals = [eid, epic_name[eid], fid, fdesc, sid, story_desc(s),
                    s.get("pri", ""), dev, "S%d" % sprint_no, tline, MOCKUP_DISPLAY]
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.alignment = WRAP if ci in (2, 4, 6, 10) else TOP
                cell.border = BORDER
            pcell = ws.cell(row=r, column=7)
            if pcell.value in PRI_FILL:
                pcell.fill = PatternFill("solid", fgColor=PRI_FILL[pcell.value])
            mcell = ws.cell(row=r, column=11)
            mcell.hyperlink = MOCKUP_URL; mcell.font = Font(color="2563EB", underline="single")
            r += 1; rows_written += 1
ws.auto_filter.ref = "A1:K%d" % (r - 1)

# Sprint summary
ws2 = wb.create_sheet("Sprint Summary")
for i, (h, w) in enumerate([("Sprint", 9), ("Timeline", 32), ("Epics in sprint", 52), ("# Features", 12), ("# User Stories", 14)], start=1):
    c = ws2.cell(row=1, column=i, value=h); c.fill = HEAD_FILL; c.font = HEAD_FONT; c.border = BORDER
    c.alignment = Alignment(vertical="center", wrap_text=True); ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
rr = 2
for n in sorted(set(EPIC_SPRINT.values())):
    eids = [e for e, sp in EPIC_SPRINT.items() if sp == n]
    epics_txt = ", ".join("%s (%s)" % (epic_name[e], e) for e in eids)
    nfeat = sum(len(FEATURES.get(e, [])) for e in eids)
    nstory = sum(len(ep["stories"]) for ep in EPICS if ep["id"] in eids)
    for ci, v in enumerate(["S%d" % n, timeline_str(n), epics_txt, nfeat, nstory], start=1):
        cell = ws2.cell(row=rr, column=ci, value=v); cell.alignment = WRAP if ci == 3 else TOP; cell.border = BORDER
    rr += 1

# Read me
ws3 = wb.create_sheet("Read me")
notes = [
    ("SafalEvents Mobile — Sprint Planning", True), ("", False),
    ("Product", False), ("SafalEvents mobile app (Expo / React Native) — host, guest & staff, with a phone-frame web build.", False),
    ("App / preview URL", False), (MOCKUP_DISPLAY, False),
    ("(Replace the Mockup URL with your Netlify web-preview URL once the web build is deployed.)", False), ("", False),
    ("How to use", True),
    ("Sheet 'Sprint Plan' = one row per user story, grouped Epic -> Feature -> User Story.", False),
    ("Sheet 'Sprint Summary' = sprint timeline overview.", False),
    ("Use the header filters to slice by Epic, Sprint, Priority or Developer.", False), ("", False),
    ("Notes", True),
    ("'Developer Assigned' values are placeholders (round-robin) — replace with your team.", False),
    ("Sprints are 2 weeks; the schedule starts %s and is a suggested sequence by dependency." % SPRINT_START.strftime("%d %b %Y"), False),
    ("Story/epic content is the single source of truth shared with SafalEvents_Mobile_User_Stories.docx.", False),
]
ws3.column_dimensions["A"].width = 110
for i, (txt, bold) in enumerate(notes, start=1):
    c = ws3.cell(row=i, column=1, value=txt)
    c.font = Font(bold=bold, size=14 if (bold and i == 1) else 11, color="F2541B" if bold else "1F2937")
    c.alignment = Alignment(wrap_text=True, vertical="center")

wb.save(OUT)
print("Sprint rows:", rows_written, "| Features:", feat_counter, "| Sprints:", len(set(EPIC_SPRINT.values())))
print("Saved:", OUT)
